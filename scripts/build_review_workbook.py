"""Build the language-review workbook a native speaker actually opens.

    python scripts/build_review_workbook.py   # -> review/Olink_Bank_Assist_language_review.xlsx

`i18n_export.py` writes the TSVs the importer reads back. This writes the
thing a reviewer works in: four sheets, colour-coded, with the context that
decides whether a translation is merely accurate or actually right.

Both exist on purpose. The TSV is the interchange format and the workbook is
the human one — and the workbook is generated from the same tables, so it
cannot drift from what is deployed.

**Why this is in the repo now.** The first workbook was built from a throwaway
script that lived nowhere, so the file in `review/` slowly aged out of the
product: by the time the admin panel reached a hundred and ninety-eight
strings, the workbook still described eighteen. A review file nobody can
regenerate is a review file that quietly stops describing the thing under
review.

Fonts are chosen for the reviewer's machine, not ours. Nyala ships with
Windows and renders Ge'ez; Arial has no Ethiopic glyphs at all, so Amharic and
Tigrinya would arrive as boxes in the two columns that matter most.
"""

from __future__ import annotations

import re
import sys
import zipfile
from datetime import UTC, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402

from bankassist.i18n import (  # noqa: E402
    _ADMIN_STRINGS,
    _NOTES,
    _STRINGS,
    _UI_STRINGS,
    LANGUAGE_NAMES,
    SUPPORTED_LANGUAGES,
)

EPOCH = datetime(2020, 1, 1, tzinfo=UTC)

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "review" / "Olink_Bank_Assist_language_review.xlsx"

ETHIOPIC = "Nyala"
BASE = "Arial"

BRAND = "1F3A5F"
HEAD_FILL = PatternFill("solid", fgColor=BRAND)
HEAD_FONT = Font(name=BASE, bold=True, color="FFFFFF", size=11)
EDIT_FILL = PatternFill("solid", fgColor="FFF9DB")   # yellow = please edit
TODO_FILL = PatternFill("solid", fgColor="FFE3E3")   # pink  = nothing supplied
LOCK_FILL = PatternFill("solid", fgColor="F1F3F5")   # grey  = leave alone
THIN = Side(style="thin", color="D0D7DE")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

EXPECTS = ["human_request", "account_specific", "complaint", "question", "greeting"]

# Which columns hold Ge'ez, given `key, context, en, am, om, ti, so, sw, notes`.
# Safe only because sw is appended at the END of SUPPORTED_LANGUAGES rather
# than inserted mid-list — see i18n.py's own note on that. Inserting a
# language before ti/am would silently point these at the wrong column.
ETHIOPIC_COLS = {4, 6}


def style_header(ws: Worksheet, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BOX
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def pin_modified(path: Path) -> None:
    """Freeze every clock reading inside the saved file.

    This is a tracked binary, so a rebuild that changes no string must
    produce no diff — otherwise the workbook shows as modified in every
    review and a real translation change is indistinguishable from a
    rebuild.

    Two clocks leak in, and pinning only the first is not enough:

    1. openpyxl writes the current time into `docProps/core.xml` on save.
       Setting `properties.modified` beforehand does not hold — openpyxl
       overwrites it as it writes — so that entry is rewritten here.
    2. Every zip member carries its own mtime in the archive headers,
       taken from the wall clock at save. Nothing a reviewer opens shows
       it, but it is four bytes per entry and it made the file differ on
       every build regardless of content.

    Both are pinned to EPOCH, so the bytes are a function of the strings
    that went in and nothing else. `tests/test_review_workbook.py` holds
    that against the committed file.
    """
    fixed = EPOCH.strftime("%Y-%m-%dT%H:%M:%SZ")
    stamp = EPOCH.timetuple()[:6]
    with zipfile.ZipFile(path) as src:
        members = [(info, src.read(info.filename)) for info in src.infolist()]

    tmp = path.with_suffix(".tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for info, data in members:
            if info.filename == "docProps/core.xml":
                data = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>" + fixed.encode() + rb"\g<2>",
                    data,
                )
            info.date_time = stamp
            out.writestr(info, data)
    tmp.replace(path)


def read_tsv(path: Path) -> list[list[str]]:
    with path.open(encoding="utf-8-sig") as fh:
        return [ln.rstrip("\n").split("\t") for ln in fh if ln.strip()]


def string_sheet(
    wb: Workbook, title: str, table: dict[str, dict[str, str]],
    context: dict[str, str] | str, height: int,
) -> Worksheet:
    """One sheet per string table. Same shape every time, so a reviewer who
    learns the first sheet has learned all of them."""
    ws = wb.create_sheet(title)
    head = ["key", "what it is for / what to preserve"]
    head += [f"{lang} — {LANGUAGE_NAMES[lang]}" for lang in SUPPORTED_LANGUAGES]
    head.append("reviewer notes")
    ws.append(head)

    for key in table["en"]:
        note = context if isinstance(context, str) else context.get(key, "")
        ws.append([key, note] + [table[lang].get(key, "") for lang in SUPPORTED_LANGUAGES] + [""])

    for i, w in enumerate([26, 58, 44, 44, 44, 44, 44, 44, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = height
        blank = not str(ws.cell(row=r, column=4).value or "").strip()
        for c in range(1, len(head) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BOX
            cell.font = Font(name=ETHIOPIC if c in ETHIOPIC_COLS else BASE, size=11)
            if c in (1, 2):
                cell.fill = LOCK_FILL
            elif c == 3:
                # English is the source. Correcting it here does nothing, and
                # a reviewer who edits it has spent effort that gets dropped.
                cell.fill = LOCK_FILL
            elif c <= 2 + len(SUPPORTED_LANGUAGES):
                # Language columns start at 3 (after key, context). Computed
                # rather than a bare literal — a hardcoded `<= 7` here is
                # exactly what left the newest language column unstyled the
                # first time this file was extended for Swahili.
                cell.fill = TODO_FILL if blank else EDIT_FILL
    style_header(ws, len(head))
    return ws


def main() -> int:
    wb = Workbook()

    # ------------------------------------------------------ Read me first
    ws = wb.active
    assert ws is not None
    ws.title = "Read me first"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 104
    ws.column_dimensions["C"].width = 34

    def line(text: str, *, size: int = 11, bold: bool = False,
             color: str = "24292F", height: int | None = None) -> None:
        r = ws.max_row + 1
        cell = ws.cell(row=r, column=2, value=text)
        cell.font = Font(name=BASE, size=size, bold=bold, color=color)
        cell.alignment = WRAP
        if height:
            ws.row_dimensions[r].height = height

    ws["B1"] = "Olink Bank Assist — language review"
    ws["B1"].font = Font(name=BASE, size=18, bold=True, color=BRAND)
    ws.row_dimensions[1].height = 26
    line("Four sheets. Only three of them are translation.", size=12, color="57606A")
    line("")
    line("Correct the six language columns. Leave the key, the context and the English "
         "columns alone — the code looks strings up by key, and edits to the English are "
         "dropped on import.", height=44)
    line("Anything in {curly braces} is a placeholder — {bank}, {n}, {name}. Keep it exactly "
         "as written and move it wherever your grammar needs it. A translation that loses one "
         "is rejected on import, because in production it would print the literal text "
         "“{n} escalations waiting” to a bank.", height=58)
    line("")
    line("Sheet “1. What the assistant says”  —  the replies customers read.",
         bold=True)
    line("Read the “what it is for” column first: several strings are wrong in a way "
         "that is invisible from the English. ack_named is a fragment with another sentence "
         "glued onto it. related_topics must be a statement, never a question. ask_contact is "
         "always the last line of a message.", height=58)
    line("")
    line("Sheet “2. What it understands”  —  sentences, and what should happen "
         "to each.", bold=True)
    line("This is the sheet that matters most. Every language problem found in the live demos "
         "so far was a sentence the assistant failed to UNDERSTAND, not a reply worded badly "
         "— a request for a manager treated as a knowledge gap, a forgotten PIN in Afaan "
         "Oromo, a transfer to a spouse refused as if it were fraud. Reviewing the replies "
         "alone would have caught none of them.", height=72)
    line("Replace every row marked FILL IN, and add as many more as you like. More sentences "
         "is always better.", height=30)
    line("")
    line("Sheets “3. Buttons customers see” and “4. The staff panel”",
         bold=True)
    line("Short labels rather than prose. Length is the constraint: each one has to fit the "
         "same space as the English, so a translation twice as long will be cut off on a "
         "phone. Sheet 4 is read by bank staff all day — a teller, a supervisor — "
         "so house terminology matters more there than elegance.", height=58)
    line("")
    line("Three things that have caught us out — please cover them",
         bold=True, color="9A3412")
    line("1.  Write words the way a customer types them, inflections included. Amharic and "
         "Tigrinya change the FINAL character rather than adding to the end, so a word we "
         "matched in its dictionary form missed every real use. አመራር "
         "→ አመራሩ was invisible to us.", height=58)
    line("2.  Give every verb for the same idea. Afaan Oromo has three words for "
         "“forgot” — irraanfachuu, dagachuu, walaaluu — where Amharic has "
         "one. We only had the first, so two of three phrasings walked straight through a "
         "security guardrail.", height=58)
    line("3.  Include sentences that must NOT trigger, marked “question”. These "
         "matter as much as the ones that should, and they are the failure nobody can see "
         "from inside the product: a real customer gets stonewalled and nothing records it as "
         "wrong.", height=58)
    line("")
    line("Colour key", bold=True)
    line("Yellow  =  please check and correct.      Pink  =  nothing supplied yet.      "
         "Grey  =  please leave as it is.")
    line("")
    line("Where it stands today", bold=True)

    # Static counts, not COUNTIF. openpyxl writes formulas with no cached
    # value, so a formula arrives blank in any viewer that reads cached values
    # — worse than a number honestly labelled as a snapshot.
    pb = read_tsv(ROOT / "review" / "phrasebook.tsv")[1:]
    counts: dict[str, list[int]] = {}
    for row in pb:
        got = counts.setdefault(row[0].strip(), [0, 0])
        got[0] += 1 if row[1].strip().startswith("FILL IN") else 0
        got[1] += 1

    status_row = ws.max_row + 1
    hdr = ws.cell(row=status_row, column=2, value="snapshot when this file was made")
    hdr.font = Font(name=BASE, bold=True, size=10, color="57606A")
    total_strings = len(_STRINGS["en"]) + len(_UI_STRINGS["en"]) + len(_ADMIN_STRINGS["en"])
    c = ws.cell(row=status_row, column=3, value=f"{total_strings} strings in 5 languages")
    c.font = Font(name=BASE, bold=True, size=10, color="57606A")
    for i, (code, name) in enumerate(LANGUAGE_NAMES.items()):
        todo, total = counts.get(code, [0, 0])
        r = status_row + 1 + i
        c1 = ws.cell(row=r, column=2, value=f"{name} ({code})")
        c1.font = Font(name=ETHIOPIC, size=11)
        text = (f"sheet 2: {todo} still to supply, of {total} rows" if todo
                else f"sheet 2: {total} rows supplied")
        c2 = ws.cell(row=r, column=3, value=text)
        c2.font = Font(name=BASE, size=11, bold=bool(todo),
                       color="B42318" if todo else "1A7F37")
    r = ws.max_row + 2
    c = ws.cell(row=r, column=2,
                value="For the live figure at any time, run:  python scripts/check_phrasebook.py")
    c.font = Font(name=BASE, size=10, italic=True, color="57606A")

    # ---------------------------------------------- 1. what it says
    string_sheet(wb, "1. What the assistant says", _STRINGS, _NOTES, height=74)

    # ------------------------------------------ 2. what it understands
    ws2 = wb.create_sheet("2. What it understands")
    rows = read_tsv(ROOT / "review" / "phrasebook.tsv")
    head2 = ["language", "phrase — what a customer types", "expect", "why it matters",
             "reviewer notes"]
    ws2.append(head2)
    example = ["ti", "ኣየናይ ሰብ ክዛረብ እደሊ  ← example: replace with a real sentence",
               "human_request", "example row — delete or overwrite", ""]
    ws2.append(example)
    for row in rows[1:]:
        ws2.append(((row + ["", "", "", ""])[:4]) + [""])

    for i, w in enumerate([12, 68, 20, 52, 30], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws2.max_row + 1):
        phrase = str(ws2.cell(row=r, column=2).value or "")
        todo = phrase.startswith("FILL IN")
        is_example = r == 2
        ws2.row_dimensions[r].height = 30
        for c_i in range(1, 6):
            cell = ws2.cell(row=r, column=c_i)
            cell.alignment = WRAP if c_i in (2, 4, 5) else TOP
            cell.border = BOX
            cell.font = Font(name=ETHIOPIC if c_i == 2 else BASE, size=11,
                             italic=is_example, color="8250DF" if is_example else "24292F")
            if todo:
                cell.fill = TODO_FILL
            elif c_i in (2, 3, 5):
                cell.fill = EDIT_FILL
            else:
                cell.fill = LOCK_FILL

    dv = DataValidation(type="list", formula1='"' + ",".join(EXPECTS) + '"', allow_blank=False)
    dv.error = "Pick one of: " + ", ".join(EXPECTS)
    dv.errorTitle = "Not a valid outcome"
    ws2.add_data_validation(dv)
    dv.add(f"C2:C{ws2.max_row + 200}")
    style_header(ws2, len(head2))

    # -------------------------------------- 3 and 4. the two interfaces
    string_sheet(
        wb, "3. Buttons customers see", _UI_STRINGS,
        "Button, label or placeholder in the chat window. Keep it short — it has to fit the "
        "same space as the English, on a phone.",
        height=44,
    )
    string_sheet(
        wb, "4. The staff panel", _ADMIN_STRINGS,
        "Read by bank staff, not customers. Keep it short — it has to fit the same space as "
        "the English. Use the words your branch staff already use.",
        height=44,
    )

    wb.properties.creator = "Olink Bank Assist"
    wb.properties.created = EPOCH

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    pin_modified(OUT)
    print(f"wrote {OUT}")
    print(f"  1. what the assistant says   {len(_STRINGS['en']):>4} strings")
    print(f"  2. what it understands       {len(rows) - 1:>4} phrases")
    print(f"  3. buttons customers see     {len(_UI_STRINGS['en']):>4} strings")
    print(f"  4. the staff panel           {len(_ADMIN_STRINGS['en']):>4} strings")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
